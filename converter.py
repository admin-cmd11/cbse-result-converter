import pandas as pd
import re
import os
import tempfile
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook

def convert_txt_to_excel(txt_file_path):
    with open(txt_file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    # Extract school info
    header_match = re.search(r'DATE:- (.*?)C\.B\.S\.E\. - (.*?)REGION: (.*?)PAGE:-', content)
    school_match = re.search(r'SCHOOL : - (\d+)\s+([^\n]+)', content)

    school_info = ""
    if header_match:
        date, exam_info, region = header_match.groups()
        school_info += f"Date: {date.strip()} | Exam: {exam_info.strip()} | Region: {region.strip()}\n"
    if school_match:
        school_code, school_name = school_match.groups()
        school_info += f"School Code: {school_code} | School Name: {school_name.strip()}"

    # Extract student records
    students = []
    records = re.finditer(
        r'(\d{8})\s+([MF])\s+([^\n]+?)\s+((?:\d{3}\s+)+)\s+(PASS|FAIL)\s+.*?\n\s*((?:\d{3}\s+[A-Z]\d\s+)+)',
        content,
        re.DOTALL
    )

    for record in records:
        roll_no, gender, name, subject_codes, result, grades_str = record.groups()
        name = ' '.join(name.split())
        subject_codes = subject_codes.strip().split()
        grade_pairs = re.findall(r'(\d{3})\s+([A-Z]\d)', grades_str)

        student_data = {
            'Roll No': roll_no,
            'Name': name,
            'Gender': gender,
            'Result': result
        }

        for i, (code, grade_pair) in enumerate(zip(subject_codes, grade_pairs)):
            if i < len(grade_pairs):
                marks, grade = grade_pair
                student_data[f'Sub {code} Marks'] = marks
                student_data[f'Sub {code} Grade'] = grade

        students.append(student_data)

    if not students:
        raise ValueError("No valid student records found!")

    df = pd.DataFrame(students)

    # Reorder columns
    base_columns = ['Roll No', 'Name', 'Gender']
    subject_columns = []
    for col in df.columns:
        if col.startswith('Sub') and 'Marks' in col:
            sub_code = col.split()[1]
            subject_columns.extend([f'Sub {sub_code} Marks', f'Sub {sub_code} Grade'])
    columns_order = base_columns + subject_columns + ['Result']
    df = df[columns_order]

    # Save to Excel
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_output:
        writer = pd.ExcelWriter(temp_output.name, engine='openpyxl')
        df.to_excel(writer, index=False, startrow=4, header=False)

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        worksheet.merge_cells('A1:Z1')
        worksheet.merge_cells('A2:Z2')

        header_cell = worksheet['A1']
        header_cell.value = school_info.split('\n')[0]
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

        school_cell = worksheet['A2']
        school_cell.value = school_info.split('\n')[1] if '\n' in school_info else ""
        school_cell.font = Font(bold=True)
        school_cell.alignment = Alignment(horizontal='center')

        for col_num, value in enumerate(df.columns.values, 1):
            cell = worksheet.cell(row=4, column=col_num, value=value)
            cell.font = Font(bold=True)

        worksheet.freeze_panes = 'A5'
        writer.close()

        return temp_output.name
